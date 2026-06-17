# """
# APS Model Answer Evaluator
# ==========================
# Compares results_*_APS_detail.json files from Gemma, Llama, Mistral
# against ground truth in ars_perceive_sub_answers_only.json

# EXPECTED STRUCTURES:
# --------------------
# results_*_APS_detail.json (model output):
# [
#   {
#     "problem_index": 1,
#     "judgments": [
#       { "ars_id": "Q1", "answer": "null", "reasoning": "..." },
#       { "ars_id": "Q2", "answer": "60_deg", "reasoning": "..." },
#       ...
#     ]
#   },
#   ...
# ]

# ars_perceive_sub_answers_only.json (ground truth):
# {
#   "1": { "Q1": "some_answer", "Q2": "60_deg", ... },
#   ...
# }
# OR list form:
# [
#   { "problem_index": 1, "Q1": "...", "Q2": "..." },
#   ...
# ]

# questions.json (optional):
# {
#   "1": { "Q1": "What is angle A?", "Q2": "What is angle B?" },
#   ...
# }

# FOLDER STRUCTURE:
# -----------------
# your_data_folder/
# ├── ars_perceive_sub_answers_only.json
# ├── questions.json                       (optional)
# ├── Gemma/
# │   ├── run1/  results_*_APS_detail.json  (up to 20 files)
# │   ├── run2/  ...
# │   └── run3/  ...
# ├── Llama/
# │   └── run1/, run2/, run3/ ...
# └── Mistral/
#     └── run1/, run2/, run3/ ...

# OUTPUT (in evaluation_results/):
# ---------------------------------
#   Gemma_run1_comparison.xlsx   ← green=correct, red=wrong, per Q
#   Llama_run2_comparison.xlsx
#   ...
#   summary_all_models.xlsx      ← accuracy table across all models/runs
#   summary_report.txt

# USAGE:
# ------
#   pip install anthropic openpyxl

#   # With Claude semantic comparison (recommended):
#   export ANTHROPIC_API_KEY="sk-ant-..."
#   python evaluate_model_answers.py --data_dir ./your_data_folder

#   # Rule-based only (no API key needed):
#   python evaluate_model_answers.py --data_dir ./your_data_folder --no_api

#   # Evaluate only selected settings for every model:
#   python evaluate_model_answers.py --settings FTFD FTNDBGC

#   # Evaluate different settings for different models:
#   python evaluate_model_answers.py \
#       --model_settings Gemma:FTFD,FTNDBGC Llama:FTFD Mistral:HTHD

#   # Custom filenames:
#   python evaluate_model_answers.py --data_dir ./data \
#       --gt_file ars_perceive_sub_answers_only.json \
#       --q_file  questions.json \
#       --out_dir evaluation_results
# """

import os, json, glob, re, argparse, time
import urllib.error
import urllib.request
from pathlib import Path
from collections import defaultdict

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False
    print("WARNING: anthropic not installed. pip install anthropic")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. pip install openpyxl  (will fallback to CSV)")

# ── CONFIG ─────────────────────────────────────────────────────`──────────────
API_KEY    = "YOUR_API_KEY"  
MODEL      = "anthropic/claude-sonnet-4-5"
MAX_TOKENS = 150
API_DELAY  = 0.15   # seconds between API calls to avoid rate limits
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

BASE_DIR = Path("/home/anshu/W/anshus_data_04-04-26/JIO/VLM Project/Testing")
DEFAULT_DATA_DIR = BASE_DIR / "APS_answers_only"
DEFAULT_GT_FILE = BASE_DIR / "ground_truth_triples" / "ars_perceive_sub_answers_only.json"
DEFAULT_Q_FILE = BASE_DIR / "ground_truth_triples" / "aps_perceive_questions_trial_test_questions_only.json"
DEFAULT_OUT_DIR = BASE_DIR / "evaluation_results"
RESULT_FILE_RE = re.compile(r"^results_(.+)_APS_(?:answers_only|detail)\.json$")

# Choose settings here instead of typing them in the terminal.
# Applies to every model. Leave empty to evaluate all settings.
# Example: SELECTED_SETTINGS = ["FTFD", "FTNDBGC"]
SELECTED_SETTINGS = ["FTFD", "FTNDBGC","FTNDBLUR","FTNDILL","FTNDIRR","FTNDPN","HTHD","HTNHBGC","HTNHBLUR","HTNHILL","HTNHIRR",
                    "HTNHPN","NHTNHD","NHTPHD","NHTSHD","NHTTHD","NTNFD","NTPFD","NTSFD","NTTFD"]

# Optional: choose different settings for each model.
# If a model is listed here, this overrides SELECTED_SETTINGS for that model.
# Example:
# MODEL_SETTINGS = {
#     "Gemma": ["FTFD", "FTNDBGC"],
#     "Llama": ["FTFD"],
#     "Mistral": ["HTHD"],
# }
MODEL_SETTINGS = {}

# Same APS problem subset used by calculate_triples_aps_answer_accuracy.py.
# Leave empty to score every problem in the questions file.
PROBLEM_INDICES = [
    1, 2, 3, 4, 7, 8, 9, 10, 11,
    14, 16, 17, 20, 22, 25, 26, 29, 30, 32, 34,
    36, 37, 38, 40, 43, 44, 45, 46, 48, 49, 50, 53, 55, 56, 57, 58, 59, 60,
    64, 72, 73, 78, 79, 81, 84, 86, 88, 95, 99, 108, 109, 112, 114, 115,
    117, 119, 120, 123, 128, 132, 133, 149, 150, 151, 152, 153, 157, 165,
    198, 207, 210, 211, 223, 239, 243, 257, 264, 281, 327, 348, 364, 365,
    420, 432, 438, 441, 447, 449, 457, 472,
]

GREEN  = PatternFill("solid", fgColor="C6EFCE") if HAS_OPENPYXL else None
RED    = PatternFill("solid", fgColor="FFC7CE") if HAS_OPENPYXL else None
BLUE   = PatternFill("solid", fgColor="4472C4") if HAS_OPENPYXL else None
NAVY   = PatternFill("solid", fgColor="1F4E79") if HAS_OPENPYXL else None
YELLOW = PatternFill("solid", fgColor="FFEB9C") if HAS_OPENPYXL else None
# ─────────────────────────────────────────────────────────────────────────────


# ── FILE DISCOVERY ────────────────────────────────────────────────────────────

def find_result_files(data_dir: Path) -> list[dict]:
    """
    Recursively finds all APS model answer JSON files.
    Returns list of {model, run, filepath}.
    """
    found = []
    MODEL_NAMES = ["Gemma", "Llama", "Mistral"]

    patterns = ["*APS_answers_only*.json", "*APS_detail*.json"]
    json_files = sorted({path for pattern in patterns for path in data_dir.rglob(pattern)})

    for json_file in json_files:
        rel = json_file.relative_to(data_dir)
        parts = rel.parts  # e.g. ('Gemma', 'run1', 'results_001_APS_detail.json')

        model_name = "Unknown"
        run_label  = "run_unknown"

        for part in parts[:-1]:
            for m in MODEL_NAMES:
                if m.lower() in part.lower():
                    model_name = m
            if re.search(r"run", part, re.IGNORECASE):
                run_label = part

        found.append({"model": model_name, "run": run_label, "filepath": json_file})

    return found


def find_file(data_dir: Path, filename: str) -> Path | None:
    direct_path = Path(filename)
    if direct_path.exists():
        return direct_path

    matches = list(data_dir.rglob(filename))
    return matches[0] if matches else None


def parse_setting(result_path: Path) -> str:
    match = RESULT_FILE_RE.match(result_path.name)
    return match.group(1) if match else result_path.stem


def sorted_question_ids(question_ids) -> list[str]:
    def key(qid: str) -> tuple[int, str]:
        match = re.search(r"\d+", str(qid))
        return (int(match.group(0)) if match else 10**9, str(qid))

    return sorted((str(qid) for qid in question_ids), key=key)


def safe_path_part(value: str) -> str:
    return re.sub(r"[^\w]+", "_", value.strip().lower()).strip("_")


def normalize_run_dir(run: str) -> str:
    run_text = run.strip().lower()
    word_to_number = {"first": "1", "second": "2", "third": "3"}
    for word, number in word_to_number.items():
        if word in run_text:
            return f"run{number}"

    match = re.search(r"run\D*(\d+)", run_text)
    if match:
        return f"run{match.group(1)}"

    return safe_path_part(run)


def comparison_output_path(out_dir: Path, model: str, run: str, setting: str) -> Path:
    return (
        out_dir
        / safe_path_part(model)
        / normalize_run_dir(run)
        / setting
        / f"{setting}_compare_eval.json"
    )


def problem_stats_output_path(out_dir: Path, model: str, run: str, setting: str) -> Path:
    return (
        out_dir
        / safe_path_part(model)
        / normalize_run_dir(run)
        / setting
        / f"{setting}_problem_stats.txt"
    )


def existing_outputs_complete(out_dir: Path, model: str, run: str, setting: str) -> bool:
    return (
        comparison_output_path(out_dir, model, run, setting).exists()
        and problem_stats_output_path(out_dir, model, run, setting).exists()
    )


def summary_from_existing_stats(
    stats_path: Path,
    model: str,
    run: str,
    setting: str,
    filename: str,
) -> dict | None:
    try:
        lines = stats_path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return None

    total_line = next((line for line in reversed(lines) if line.strip().startswith("TOTAL")), "")
    match = re.search(r"(\d+)\s*/\s*(\d+)\s*=\s*([0-9]+(?:\.[0-9]+)?)%", total_line)
    if not match:
        return None

    correct = int(match.group(1))
    total = int(match.group(2))
    accuracy = float(match.group(3))
    return {
        "model": model,
        "run": run,
        "setting": setting,
        "file": filename,
        "correct": correct,
        "total": total,
        "accuracy": accuracy,
    }


def normalize_setting_name(setting: str) -> str:
    return setting.strip().upper()


def parse_settings_arg(values) -> set[str]:
    if not values:
        return set()

    settings = set()
    for value in values:
        for setting in str(value).split(","):
            setting = normalize_setting_name(setting)
            if setting:
                settings.add(setting)
    return settings


def parse_model_settings_arg(values) -> dict[str, set[str]]:
    if not values:
        return {}

    parsed = defaultdict(set)
    for value in values:
        if ":" in value:
            model, settings_text = value.split(":", 1)
        elif "=" in value:
            model, settings_text = value.split("=", 1)
        else:
            raise ValueError(
                f"Invalid --model_settings value '{value}'. Use Model:SETTING1,SETTING2"
            )

        model_key = model.strip().lower()
        if not model_key:
            raise ValueError(f"Missing model name in --model_settings value '{value}'")

        parsed[model_key].update(parse_settings_arg([settings_text]))

    return dict(parsed)


def result_file_allowed(
    model: str,
    setting: str,
    selected_settings: set[str],
    model_settings: dict[str, set[str]],
) -> bool:
    setting_key = normalize_setting_name(setting)
    model_key = model.strip().lower()

    if model_key in model_settings:
        return setting_key in model_settings[model_key]

    if selected_settings:
        return setting_key in selected_settings

    return True


# ── JSON LOADERS ──────────────────────────────────────────────────────────────

def load_json(path) -> any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_ground_truth(path) -> dict:
    """
    Returns: { "1": {"Q1": "answer", "Q2": "answer", ...}, ... }
    Handles both dict and list formats.
    """
    raw = load_json(path)
    result = {}

    if isinstance(raw, dict):
        for idx, val in raw.items():
            if isinstance(val, dict):
                result[str(idx)] = {str(k): str(v) for k, v in val.items()}
            elif isinstance(val, list):
                # list of {"ars_id": ..., "answer": ...}
                result[str(idx)] = {}
                for item in val:
                    qid = item.get("ars_id") or item.get("id") or item.get("question_id")
                    ans = item.get("answer") or item.get("ANSWER") or ""
                    if qid:
                        result[str(idx)][str(qid)] = str(ans)

    elif isinstance(raw, list):
        for item in raw:
            idx = str(item.get("problem_index") or item.get("index") or item.get("id", ""))
            if not idx:
                continue
            result[idx] = {}
            for k, v in item.items():
                if k not in ("problem_index", "index", "id"):
                    result[idx][str(k)] = str(v)

    return result


def load_questions(path) -> dict:
    """
    Returns: { "1": {"Q1": "question text", "Q2": "..."}, ... }
    """
    raw = load_json(path)
    result = {}

    if isinstance(raw, dict):
        for idx, val in raw.items():
            if isinstance(val, dict):
                result[str(idx)] = {str(k): str(v) for k, v in val.items()}
            elif isinstance(val, list):
                result[str(idx)] = {}
                for item in val:
                    qid = item.get("ars_id") or item.get("id") or item.get("question_id")
                    q   = item.get("question") or item.get("text") or ""
                    if qid:
                        result[str(idx)][str(qid)] = str(q)
    elif isinstance(raw, list):
        for item in raw:
            idx = str(item.get("problem_index") or item.get("index") or item.get("id", ""))
            if not idx:
                continue
            result[idx] = {}
            for k, v in item.items():
                if k.startswith("Q") or k.startswith("q"):
                    result[idx][k] = str(v)

    return result


def load_model_file(path) -> dict:
    """
    Parses model output file.
    Returns: { "1": {"Q1": {"answer": ..., "reasoning": ...}, ...}, ... }
    """
    raw = load_json(path)
    result = {}

    entries = raw if isinstance(raw, list) else raw.get("results", [raw])

    for entry in entries:
        idx = str(entry.get("problem_index") or entry.get("index") or entry.get("id", ""))
        if not idx:
            continue
        result[idx] = {}
        for j in entry.get("judgments", []):
            qid = str(j.get("ars_id") or j.get("id") or j.get("question_id", ""))
            result[idx][qid] = {
                "answer":    str(j.get("answer", "")).strip(),
                "reasoning": str(j.get("reasoning", "")).strip(),
            }

    return result


# ── COMPARISON ───────────────────────────────────────────────────────────────

_client = None


def _base_clean(value) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    if not text:
        return ""

    if text in {"none", "null", "nan", "n/a", "unknown", "not provided", "not specified"}:
        return ""

    replacements = {
        "_": " ",
        "-": " ",
        "∠": " angle ",
        "\\angle": " angle ",
        "\\deg": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r"\b(the|a|an)\b", "", text)
    text = re.sub(r"\blies\s+on\b", "on", text)
    text = re.sub(r"\bis\s+on\b", "on", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def numeric_value(value) -> float | None:
    text = _base_clean(value)
    if not text:
        return None

    text = text.replace("°", "")
    text = re.sub(r"\b(degrees?|deg|units?|cm|mm|m|km|meters?)\b", "", text)
    text = text.strip()

    equation = re.fullmatch(r"[a-z]\s*=\s*([-+]?\d+(?:\.\d+)?)", text)
    if equation:
        return float(equation.group(1))

    match = re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def canonical_answer(value) -> str:
    text = _base_clean(value)
    if not text:
        return ""

    number = numeric_value(text)
    if number is not None:
        if abs(number - 90.0) < 1e-6:
            return "90"
        return f"{number:g}"

    compact = re.sub(r"[\s,]+", "", text)

    if compact in {"rightangle", "rightangled", "right"}:
        return "90"

    if compact in {"parallel", "isparallel", "parallelto", "||"}:
        return "parallel"

    if compact in {"perpendicular", "isperpendicular", "perpendicularto", "⟂", "perp"}:
        return "perpendicular"

    if compact in {"on", "lieson", "ison"}:
        return "on"

    if compact in {"bisects", "bisect", "isbisectorof", "anglebisector", "bisector"}:
        return "bisects"

    text = text.replace("°", "")
    text = re.sub(r"\b(degrees?|deg|units?|cm|mm|m|km|meters?)\b", "", text)
    text = re.sub(r"\b(point|line|side|angle)\b", "", text)
    text = re.sub(r"[^a-z0-9.,/ ]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def token_set(value) -> set[str]:
    return {tok for tok in canonical_answer(value).split() if tok}


def answers_match(predicted, gold) -> bool:
    pred_clean = canonical_answer(predicted)
    gold_clean = canonical_answer(gold)
    if not pred_clean or not gold_clean:
        return False

    if pred_clean == gold_clean:
        return True

    pred_num = numeric_value(predicted)
    gold_num = numeric_value(gold)
    if pred_num is not None and gold_num is not None:
        return abs(pred_num - gold_num) < 1e-6

    pred_tokens = token_set(pred_clean)
    gold_tokens = token_set(gold_clean)
    if gold_tokens and gold_tokens.issubset(pred_tokens):
        return True

    pred_compact = re.sub(r"[\s,]+", "", pred_clean)
    gold_compact = re.sub(r"[\s,]+", "", gold_clean)
    return bool(gold_compact and pred_compact == gold_compact)


def get_client():
    global _client
    if _client is None and HAS_ANTHROPIC and API_KEY and not is_openrouter_key():
        _client = anthropic.Anthropic(api_key=API_KEY)
    return _client


def is_openrouter_key() -> bool:
    return API_KEY.startswith("sk-or-")


def has_api_client() -> bool:
    return bool(API_KEY) and (is_openrouter_key() or HAS_ANTHROPIC)


def call_openrouter(prompt: str) -> str:
    payload = {
        "model": MODEL,
        "max_tokens": MAX_TOKENS,
        "messages": [{"role": "user", "content": prompt}],
    }
    request = urllib.request.Request(
        OPENROUTER_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "http://localhost",
            "X-Title": "APS Answer Evaluator",
        },
        method="POST",
    )

    with urllib.request.urlopen(request, timeout=60) as response:
        data = json.loads(response.read().decode("utf-8"))

    return data["choices"][0]["message"]["content"]


def semantic_compare(question: str, ground_truth: str, model_answer: str, reasoning: str) -> tuple[bool, str]:
    """
    Uses Claude API for semantic comparison.
    KEY RULE: if final answer is wrong vs GT, it's WRONG — even if reasoning is correct.
    Falls back to rule_based if API unavailable.
    """
    if not has_api_client():
        return rule_based_compare(ground_truth, model_answer)

    prompt = f"""You are evaluating a visual question answering model's output.

Question: {question if question else "(not provided)"}
Ground Truth Answer: {ground_truth}
Model's Final Answer: {model_answer}
Model's Reasoning: {reasoning}

Evaluation rules :
1. Compare only the "Model's Final Answer" against "Ground Truth Answer".
2. Even if the reasoning contains the correct answer, if the final answer field is wrong → mark wrong.
3. Semantically equivalent answers are correct:
   - "Yes" is correct if GT is a property confirming the question (e.g., Q: "is O the center?" GT: "center" → "Yes" is correct)
   - Unit variations: "60_deg" = "60°" = "60 degrees"
   - Phrasing: "lies_on AB" = "on AB" = "AB"
   - "null" or "none" = "not visible" = "cannot determine" (these match each other)
4. Partially correct or vague answers are wrong.
5. Respond ONLY with valid JSON, no extra text:
{{"correct": true or false, "reason": "one short sentence"}}"""

    try:
        if is_openrouter_key():
            text = call_openrouter(prompt)
        else:
            response = get_client().messages.create(
                model=MODEL, max_tokens=MAX_TOKENS,
                messages=[{"role": "user", "content": prompt}]
            )
            text = response.content[0].text

        text = re.sub(r"```json|```", "", text).strip()
        data = json.loads(text)
        return bool(data.get("correct", False)), data.get("reason", "")
    except (urllib.error.URLError, KeyError, json.JSONDecodeError, Exception) as exc:
        return rule_based_compare(ground_truth, model_answer)


def rule_based_compare(ground_truth: str, model_answer: str) -> tuple[bool, str]:
    if answers_match(model_answer, ground_truth):
        return True, "Canonical answer match"

    return False, "Answer mismatch"


def compare_entry(problem_idx: str, qid: str,
                  question: str, ground_truth: str,
                  model_ans: str, reasoning: str,
                  use_api: bool) -> dict:

    if model_ans in ("", "MISSING") or canonical_answer(model_ans) == "":
        return {"correct": False, "reason": "Model answer missing"}
    if ground_truth in ("", "MISSING"):
        return {"correct": False, "reason": "Ground truth missing"}

    if use_api:
        correct, reason = semantic_compare(question, ground_truth, model_ans, reasoning)
        time.sleep(API_DELAY)
    else:
        correct, reason = rule_based_compare(ground_truth, model_ans)

    return {"correct": correct, "reason": reason}


# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────

def write_comparison_excel(rows: list[dict], output_path: str, title: str) -> tuple[int, int]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparisons"

    headers     = ["Problem Index", "Q ID", "Question", "Ground Truth", "Model Answer", "Reasoning", "Result", "Notes"]
    col_widths  = [14, 8, 45, 22, 22, 45, 12, 30]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    tc = ws.cell(row=1, column=1, value=title)
    tc.fill = NAVY
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = BLUE
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 18

    correct_count = 0
    for r, row in enumerate(rows, start=3):
        fill = GREEN if row["correct"] else RED
        vals = [
            row["problem_index"],
            row["qid"],
            row["question"],
            row["ground_truth"],
            row["model_answer"],
            row["reasoning"],
            "✓ CORRECT" if row["correct"] else "✗ WRONG",
            row["reason"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 45
        if row["correct"]:
            correct_count += 1

    total = len(rows)
    acc   = correct_count / total * 100 if total else 0

    # Summary footer
    sr = total + 3
    ws.cell(row=sr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=sr, column=2,
            value=f"Correct: {correct_count} / {total}  →  Accuracy: {acc:.1f}%").font = Font(bold=True)

    ws.freeze_panes = "A3"
    wb.save(output_path)
    return correct_count, total


def write_summary_excel(summaries: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = ["Model", "Run", "File", "Correct", "Total", "Accuracy (%)"]
    widths  = [12, 14, 50, 10, 10, 16]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    tc = ws.cell(row=1, column=1, value="APS Evaluation — All Models Summary")
    tc.fill = NAVY
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.alignment = Alignment(horizontal="center")

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = BLUE
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    for r, s in enumerate(summaries, start=3):
        ws.cell(row=r, column=1, value=s["model"])
        ws.cell(row=r, column=2, value=s["run"])
        ws.cell(row=r, column=3, value=s["file"])
        ws.cell(row=r, column=4, value=s["correct"])
        ws.cell(row=r, column=5, value=s["total"])
        acc_cell = ws.cell(row=r, column=6, value=round(s["accuracy"], 1))
        acc_cell.fill = GREEN if s["accuracy"] >= 70 else (YELLOW if s["accuracy"] >= 50 else RED)

    wb.save(output_path)


def write_comparison_json(rows: list[dict], output_path: Path) -> tuple[int, int]:
    grouped = defaultdict(list)
    correct_count = 0

    for row in rows:
        if row["correct"]:
            correct_count += 1

        grouped[str(row["problem_index"])].append(
            {
                "qid": row["qid"],
                "question": row["question"],
                "GT_answer": row["ground_truth"],
                "model_answer": row["model_answer"],
                "correct/incorrect": "correct" if row["correct"] else "incorrect",
            }
        )

    def problem_key(problem_index: str):
        return int(problem_index) if problem_index.isdigit() else problem_index

    output = {
        problem_index: grouped[problem_index]
        for problem_index in sorted(grouped, key=problem_key)
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
        f.write("\n")

    return correct_count, len(rows)


def write_problem_stats_table(rows: list[dict], output_path: Path) -> Path:
    stats = defaultdict(lambda: {"aps_questions": 0, "answered": 0, "correct": 0})

    for row in rows:
        problem_index = str(row["problem_index"])
        model_answer = row["model_answer"]

        stats[problem_index]["aps_questions"] += 1
        if model_answer not in ("", "MISSING") and canonical_answer(model_answer) != "":
            stats[problem_index]["answered"] += 1
        if row["correct"]:
            stats[problem_index]["correct"] += 1

    def problem_key(problem_index: str):
        return int(problem_index) if problem_index.isdigit() else problem_index

    lines = [
        "Per-problem table:",
        (
            f"{'Problem_Index':<15} | {'APS questions':<13} | {'Answered':<8} | "
            f"{'Correct':<7} | {'Accuracy':<18}"
        ),
        "-" * 75,
    ]

    total_questions = 0
    total_answered = 0
    total_correct = 0

    for problem_index in sorted(stats, key=problem_key):
        problem_stats = stats[problem_index]
        aps_questions = problem_stats["aps_questions"]
        answered = problem_stats["answered"]
        correct = problem_stats["correct"]
        accuracy = correct / aps_questions * 100 if aps_questions else 0

        total_questions += aps_questions
        total_answered += answered
        total_correct += correct

        lines.append(
            f"{problem_index:<15} | {aps_questions:<13} | {answered:<8} | "
            f"{correct:<7} | {correct}/{aps_questions} = {accuracy:.2f}%"
        )

    total_accuracy = total_correct / total_questions * 100 if total_questions else 0
    lines.extend(
        [
            "-" * 75,
            (
                f"{'TOTAL':<15} | {total_questions:<13} | {total_answered:<8} | "
                f"{total_correct:<7} | {total_correct}/{total_questions} = {total_accuracy:.2f}%"
            ),
        ]
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output_path


def build_overall_model_accuracy_report(summaries: list[dict]) -> str:
    model_totals = defaultdict(lambda: {"correct": 0, "total": 0, "files": 0, "runs": set()})
    run_totals = defaultdict(lambda: {"correct": 0, "total": 0, "files": 0})
    model_setting_totals = defaultdict(lambda: {"correct": 0, "total": 0, "files": 0})
    overall_setting_totals = defaultdict(lambda: {"correct": 0, "total": 0, "files": 0})

    for summary in summaries:
        model = summary["model"]
        run = summary["run"]
        setting = summary["setting"]
        correct = summary["correct"]
        total = summary["total"]

        model_totals[model]["correct"] += correct
        model_totals[model]["total"] += total
        model_totals[model]["files"] += 1
        model_totals[model]["runs"].add(run)

        run_key = (model, run)
        run_totals[run_key]["correct"] += correct
        run_totals[run_key]["total"] += total
        run_totals[run_key]["files"] += 1

        model_setting_key = (model, setting)
        model_setting_totals[model_setting_key]["correct"] += correct
        model_setting_totals[model_setting_key]["total"] += total
        model_setting_totals[model_setting_key]["files"] += 1

        overall_setting_totals[setting]["correct"] += correct
        overall_setting_totals[setting]["total"] += total
        overall_setting_totals[setting]["files"] += 1

    lines = [
        "APS Answer Accuracy Summary",
        "=" * 65,
        "Accuracy calculation follows calculate_triples_aps_answer_accuracy.py:",
        "- denominator comes from the questions JSON",
        "- missing model answers are counted as incorrect",
        "- selected APS problem indices are used",
        "- standard deviation and triples-used metrics are omitted",
        "",
        "Per-model metrics:",
        f"{'Model':<12} | {'Avg run acc':<12} | {'Pooled accuracy':<20} | {'Runs':<4} | {'Settings':<8}",
        "-" * 75,
    ]

    for model in sorted(model_totals):
        totals = model_totals[model]
        run_accuracies = []
        for run in sorted(totals["runs"]):
            run_total = run_totals[(model, run)]
            run_accuracy = run_total["correct"] / run_total["total"] if run_total["total"] else 0
            run_accuracies.append(run_accuracy)

        avg_run_accuracy = sum(run_accuracies) / len(run_accuracies) * 100 if run_accuracies else 0
        pooled_accuracy = totals["correct"] / totals["total"] * 100 if totals["total"] else 0
        lines.append(
            f"{model:<12} | {avg_run_accuracy:>9.2f}% | "
            f"{totals['correct']:5}/{totals['total']:<5} = {pooled_accuracy:6.2f}% | "
            f"{len(totals['runs']):<4} | {totals['files']:<8}"
        )

    lines.extend(["", "Per-run metrics:", f"{'Model':<12} | {'Run':<20} | {'Accuracy':<20} | {'Settings':<8}", "-" * 70])

    for model, run in sorted(run_totals):
        totals = run_totals[(model, run)]
        accuracy = totals["correct"] / totals["total"] * 100 if totals["total"] else 0
        lines.append(
            f"{model:<12} | {run:<20} | {totals['correct']:5}/{totals['total']:<5} = {accuracy:6.2f}% | "
            f"{totals['files']:<8}"
        )

    lines.extend(
        [
            "",
            "Case metrics combined across all runs, per model:",
            f"{'Model':<12} | {'Case':<12} | {'Accuracy':<20} | {'Result files':<12}",
            "-" * 72,
        ]
    )

    for model, setting in sorted(model_setting_totals):
        totals = model_setting_totals[(model, setting)]
        accuracy = totals["correct"] / totals["total"] * 100 if totals["total"] else 0
        lines.append(
            f"{model:<12} | {setting:<12} | {totals['correct']:5}/{totals['total']:<5} = {accuracy:6.2f}% | "
            f"{totals['files']:<12}"
        )

    lines.extend(
        [
            "",
            "Case metrics combined across all models and runs:",
            f"{'Case':<12} | {'Accuracy':<20} | {'Result files':<12}",
            "-" * 52,
        ]
    )

    for setting in sorted(overall_setting_totals):
        totals = overall_setting_totals[setting]
        accuracy = totals["correct"] / totals["total"] * 100 if totals["total"] else 0
        lines.append(
            f"{setting:<12} | {totals['correct']:5}/{totals['total']:<5} = {accuracy:6.2f}% | "
            f"{totals['files']:<12}"
        )

    return "\n".join(lines)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data_dir",  default=str(DEFAULT_DATA_DIR))
    parser.add_argument("--gt_file",   default=str(DEFAULT_GT_FILE))
    parser.add_argument("--q_file",    default=str(DEFAULT_Q_FILE))
    parser.add_argument("--out_dir",   default=str(DEFAULT_OUT_DIR))
    parser.add_argument(
        "--settings",
        nargs="*",
        default=None,
        help="Evaluate only these settings for every model, e.g. --settings FTFD FTNDBGC",
    )
    parser.add_argument(
        "--model_settings",
        nargs="*",
        default=None,
        help=(
            "Evaluate model-specific settings, e.g. "
            "--model_settings Gemma:FTFD,FTNDBGC Llama:FTFD Mistral:HTHD"
        ),
    )
    parser.add_argument("--no_api",    action="store_true", help="Use rule-based comparison only")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    out_dir  = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    try:
        selected_settings = (
            parse_settings_arg(args.settings)
            if args.settings is not None
            else parse_settings_arg(SELECTED_SETTINGS)
        )
        model_settings = (
            parse_model_settings_arg(args.model_settings)
            if args.model_settings is not None
            else {
                model.strip().lower(): parse_settings_arg(settings)
                for model, settings in MODEL_SETTINGS.items()
            }
        )
    except ValueError as exc:
        print(f"ERROR: {exc}")
        return

    use_api = has_api_client() and not args.no_api
    comparison_mode = "Model semantic (OpenRouter API)" if is_openrouter_key() else "Claude semantic (Anthropic API)"

    print(f"\n{'='*65}")
    print(f"  APS Model Answer Evaluator")
    print(f"{'='*65}")
    print(f"  Data dir    : {data_dir}")
    print(f"  Output dir  : {out_dir}")
    print(f"  Comparison  : {comparison_mode if use_api else 'Rule-based (no API)'}")
    if selected_settings:
        print(f"  Settings    : {', '.join(sorted(selected_settings))}")
    if model_settings:
        model_settings_text = "; ".join(
            f"{model}: {', '.join(sorted(settings))}"
            for model, settings in sorted(model_settings.items())
        )
        print(f"  Model sets  : {model_settings_text}")
    print(f"{'='*65}\n")

    # Load ground truth
    gt_path = find_file(data_dir, args.gt_file)
    if not gt_path:
        print(f"ERROR: Cannot find {args.gt_file} in {data_dir}"); return
    print(f"Ground truth : {gt_path}")
    gt_data = load_ground_truth(str(gt_path))
    print(f"             → {len(gt_data)} problem indices loaded")

    # Load questions (optional)
    q_data = {}
    if args.q_file:
        q_path = find_file(data_dir, args.q_file)
        if q_path:
            print(f"Questions    : {q_path}")
            q_data = load_questions(str(q_path))

    # Find all model result files
    print(f"\nScanning for model result files...")
    result_files = find_result_files(data_dir)
    if not result_files:
        print("ERROR: No *APS_answers_only*.json or *APS_detail*.json files found. Check folder structure."); return

    before_filter_count = len(result_files)
    result_files = [
        rf for rf in result_files
        if result_file_allowed(
            rf["model"],
            parse_setting(rf["filepath"]),
            selected_settings,
            model_settings,
        )
    ]
    if not result_files:
        print("ERROR: No result files matched the selected settings."); return

    if len(result_files) != before_filter_count:
        print(f"Found {before_filter_count} result files; evaluating {len(result_files)} after settings filter.\n")
    else:
        print(f"Found {len(result_files)} result files.\n")

    summaries = []

    for rf in result_files:
        model    = rf["model"]
        run      = rf["run"]
        filepath = rf["filepath"]
        setting  = parse_setting(filepath)

        print(f"── {model} / {run} / {filepath.name}")

        out_path = comparison_output_path(out_dir, model, run, setting)
        stats_path = problem_stats_output_path(out_dir, model, run, setting)
        if existing_outputs_complete(out_dir, model, run, setting):
            existing_summary = summary_from_existing_stats(
                stats_path,
                model,
                run,
                setting,
                filepath.name,
            )
            if existing_summary:
                summaries.append(existing_summary)
                print(f"   SKIP existing output → {out_path}")
                print(f"   Existing stats       → {stats_path}")
                continue

            print("   Existing output found but stats could not be parsed; regenerating.")

        try:
            model_data = load_model_file(str(filepath))
        except Exception as e:
            print(f"   ERROR loading: {e}"); continue

        # Build comparison rows
        rows = []
        selected_problem_indices = set(PROBLEM_INDICES) if PROBLEM_INDICES else None
        denominator_data = q_data if q_data else gt_data
        all_indices = []

        for idx in denominator_data.keys():
            try:
                idx_int = int(idx)
            except (TypeError, ValueError):
                continue

            if selected_problem_indices and idx_int not in selected_problem_indices:
                continue
            all_indices.append(str(idx))

        all_indices = sorted(all_indices, key=lambda x: int(x) if x.isdigit() else x)

        for idx in all_indices:
            m_entry  = model_data.get(idx, {})
            gt_entry = gt_data.get(idx, {})
            q_entry  = q_data.get(idx, {})

            qid_source = q_entry if q_entry else gt_entry
            all_qids = sorted_question_ids(qid_source.keys())

            for qid in all_qids:
                m_qa       = m_entry.get(qid, {})
                model_ans  = m_qa.get("answer", "MISSING") if m_qa else "MISSING"
                reasoning  = m_qa.get("reasoning", "") if m_qa else ""
                gt_ans     = gt_entry.get(qid, "MISSING")
                question   = q_entry.get(qid, "")

                result = compare_entry(idx, qid, question, gt_ans, model_ans, reasoning, use_api)

                rows.append({
                    "problem_index": idx,
                    "qid":           qid,
                    "question":      question,
                    "ground_truth":  gt_ans,
                    "model_answer":  model_ans,
                    "reasoning":     reasoning,
                    "correct":       result["correct"],
                    "reason":        result["reason"],
                })

        # Write per-setting JSON output.
        correct, total = write_comparison_json(rows, out_path)
        write_problem_stats_table(rows, stats_path)

        acc = correct / total * 100 if total else 0
        print(f"   {correct}/{total} correct = {acc:.1f}%  →  {out_path}")
        print(f"   Per-problem stats → {stats_path}")

        summaries.append({
            "model": model, "run": run,
            "setting": setting,
            "file": filepath.name,
            "correct": correct, "total": total, "accuracy": acc
        })

    # Summary files
    print(f"\n{'='*65}")
    print("  FINAL SUMMARY")
    print(f"{'='*65}")
    lines = ["APS Evaluation Summary", "="*65]
    for s in summaries:
        line = f"{s['model']:8} | {s['run']:12} | {s['file'][:40]:40} | {s['correct']:4}/{s['total']:4} = {s['accuracy']:5.1f}%"
        print(line); lines.append(line)

    (out_dir / "summary_report.txt").write_text("\n".join(lines), encoding="utf-8")

    overall_report = build_overall_model_accuracy_report(summaries)
    overall_path = out_dir / "overall_model_accuracy.txt"
    overall_path.write_text(overall_report, encoding="utf-8")
    print(f"\nOverall model accuracy saved → {overall_path}")

    if HAS_OPENPYXL and summaries:
        write_summary_excel(summaries, str(out_dir / "summary_all_models.xlsx"))
        print(f"\nSummary saved → {out_dir / 'summary_all_models.xlsx'}")

    print(f"All files in  → {out_dir}/\nDone ✓\n")


if __name__ == "__main__":
    main()